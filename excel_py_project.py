import openpyxl

#setting a path
path="student_reccords.xlsx"
wb=openpyxl.load_workbook(filename=path)
#creating a file
wb.save(filename=path)
sheet=wb.active
m_row=sheet.max_row
for row in sheet.iter_rows(min_row=2,max_row=m_row):
    row[6].value = row[3].value+row[4].value+row[5].value
while True:
    print("1. To add data\n2. To remove data\n3. To update data\n4. To print your sheet\n5. To search for student details ")
    ch=int(input("enter your choice : "))
    if ch ==1:
        row=int(input("enter the row number in which you want to enter data : "))
        column=int(input("enter the column number in which you want to enter data : "))
        data=input("enter the data : ")
        sheet.cell(row=row, column=column).value=data
        wb.save(filename=path)
    elif ch==2:
        row=int(input("enter the row number in which you want to remove data from : "))
        column=int(input("enter the column number in which you want to remove data from  : "))
        sheet.cell(row=row, column=column).value=None
        wb.save(filename=path)
    elif ch==3:
        row=int(input("enter the row number in which you want to enter data : "))
        column=int(input("enter the column number in which you want to enter data : "))
        data=input("enter the data : ")
        sheet.cell(row=row, column=column).value=data
        wb.save(filename=path)
    elif ch==4:
        row=sheet.max_row
        column=sheet.max_column
        for i in range(1,row+1):
            for j in range (1,column+1):
                cell=sheet.cell(row=i,column=j).value
                if cell ==None:
                    continue
                else:
                    print( cell, end="  ")
            print("\n")
    elif ch==5:
        x=int(input("Enter the total marks of the student you want to search : "))
        for row in sheet.iter_rows(min_row=1,max_row=1):
            for cell in row:
                print (cell.value,end=" ")
        print()
        m_row=sheet.max_row
        for row in sheet.iter_rows(min_row=2,max_row=m_row):
            if row[3].value+row[4].value+row[5].value == x:
                for cell in row:
                    print(cell.value, end=" ")
                print()
        else:
            print("No student with",x,"total marks")
    else:
        print("choose from option 1 to 5")
    
     
# previous logic for ch==5  
"""
        l=[]
        row=sheet.max_row
        for i in range(2,row):
            l.append(sheet.cell(row=i,column=7).value)
        for j in range(0,len(l)):
            if l[j]==x:
                column=sheet.max_column
                for k in range (1,column+1):
                    print(sheet.cell(row=1,column=k).value,end="    ")
                print("\n")
                for k in range (1,column+1):
                    if sheet.cell(row=i+2,column=k).value==None:
                        continue
                    else:
                        print(sheet.cell(row=i+2,column=k).value,end="    ")
            else:
                print("No student with" ,x,"total marks")
                break
"""
